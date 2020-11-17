from flask import render_template,flash, redirect, url_for, Blueprint,request,send_file
from flask_login import login_user , logout_user, login_required, login_manager,current_user
import openpyxl
import os
from werkzeug.utils import secure_filename
from datetime import  datetime, date
import pandas as pd
from mailmerge import MailMerge
import mammoth


dimci = Blueprint('dimci', __name__)

@dimci.route("/index_Dimci")
def index_Dimci():
    if current_user.is_authenticated:
        return render_template('index_dimci.html', name=current_user.nome)
    return redirect(url_for('auth.login'))

#####################################################################
#####################################################################
####                    PLANILHAS      ##############################
#####################################################################

@dimci.route('/listarPlanilhas',methods = ['POST','GET'])
@login_required
def listarPlanilhas():
    DOWNLOAD_FOLDER = os.path.join(os.getcwd(), 'SGI_v0/app/static/Excel/Planilhas_Dimci')
    files = os.listdir(DOWNLOAD_FOLDER)
    return render_template('listar_planilhas.html', files=files)

@dimci.route('/uploadPlanilha',methods = ['POST'])
@login_required
def uploadPlanilha():

    if request.method == 'POST':

        UPLOAD_FOLDER = os.path.join(os.getcwd(), 'SGI_v0/app/static/Excel/Planilhas_Dimci')
        file = request.files['nome']
        data = datetime.strftime(datetime.now(),'%d_%m_%Y_%H_%M')
        #savePath = os.path.join(UPLOAD_FOLDER , secure_filename(current_user.nome+'_'+data+''+file.filename))
        savePath = os.path.join(UPLOAD_FOLDER, secure_filename(file.filename))

        file.save(savePath)
        return redirect(url_for('dimci.listarPlanilhas'))

@dimci.route('/downloadPlanilha/<file>',methods = ['POST','GET'])
@login_required
def downloadPlanilha(file):
    path = os.path.join(os.getcwd(), 'SGI_v0/app/static/Excel/Planilhas_Dimci')

    arquivo = os.path.join(path, file)

    return send_file(arquivo, mimetype='imagem/png')

@dimci.route('/excluirPlanilha/<file>',methods = ['POST','GET'])
@login_required
def excluirPlanilha(file):
    path = os.path.join(os.getcwd(), 'SGI_v0/app/static/Excel/Planilhas_Dimci')

    arquivo = os.path.join(path, file)

    if os.path.exists(arquivo):
        print('DELETAR', arquivo)
        os.remove(arquivo)

    return redirect(url_for('dimci.listarPlanilhas'))

@dimci.route('/visualizarPlanilha/<file>',methods = ['POST','GET'])
@login_required
def visualizarPlanilha(file):
    path = os.path.join(os.getcwd(), 'SGI_v0/app/static/Excel/Planilhas_Dimci')

    arquivo = os.path.join(path, file)

    html = pd.read_excel(arquivo).to_html()
    return render_template('planilha.html', html_code=html)


@dimci.route('/selecionarPlanilhas', methods = ['POST'])
@login_required
def selecionarPlanilhas():
    if request.method == 'POST':
        path = os.path.join(os.getcwd(), 'SGI_v0/app/static/Excel/Planilhas_Dimci')

        lista = request.form.getlist('selecionar')
        arquivo = os.path.join(path, lista[0])
        df = pd.read_excel(arquivo)
        if lista:
            for i in lista[1:]:
                arquivo = os.path.join(path, i)
                p = pd.read_excel(arquivo)
                print('P',p)
                df = df.append(p,ignore_index=True)
                print('df', df)

        else:
            return redirect(url_for('dimci.listarPlanilhas'))



        date = datetime.now()
        date = date.strftime("%d_%m_%Y_%H_%M_%S")

        arquivo = os.path.join(path, 'Agrupamento_{}_{}.xlsx'.format(current_user.nome,date))

        df.to_excel(arquivo, index=True, header=True)

        return redirect(url_for('dimci.listarPlanilhas'))
#
###########################################################################
#####################################################################
#####################################################################
####                    Gerenciar Formulários     ##############################
#####################################################################

@dimci.route('/listarFormularios',methods = ['POST','GET'])
@login_required
def listarFormularios():
    DOWNLOAD_FOLDER = os.path.join(os.getcwd(), 'SGI_v0/app/static/Word/formularios')
    files = os.listdir(DOWNLOAD_FOLDER)
    return render_template('listar_formularios.html', files=files)

@dimci.route('/uploadFormulario',methods = ['POST'])
@login_required
def uploadFormulario():

    if request.method == 'POST':

        UPLOAD_FOLDER = os.path.join(os.getcwd(), 'SGI_v0/app/static/Word/formularios')
        file = request.files['nome']
        data = datetime.strftime(datetime.now(),'%d_%m_%Y_%H_%M')
        #savePath = os.path.join(UPLOAD_FOLDER , secure_filename(current_user.nome+'_'+data+''+file.filename))
        savePath = os.path.join(UPLOAD_FOLDER, secure_filename(file.filename))

        file.save(savePath)
        return redirect(url_for('dimci.listarFormularios'))

@dimci.route('/downloadFormulario/<file>',methods = ['POST','GET'])
@login_required
def downloadFormulario(file):
    path = os.path.join(os.getcwd(), 'SGI_v0/app/static/Word/formularios')

    arquivo = os.path.join(path, file)

    return send_file(arquivo, mimetype='imagem/png')

@dimci.route('/excluirFormulario/<file>',methods = ['POST','GET'])
@login_required
def excluirFormulario(file):
    path = os.path.join(os.getcwd(), 'SGI_v0/app/static/Word/formularios')

    arquivo = os.path.join(path, file)

    if os.path.exists(arquivo):
        print('DELETAR', arquivo)
        os.remove(arquivo)

    return redirect(url_for('dimci.listarFormularios'))

@dimci.route('/visualizarFormulario/<file>',methods = ['POST','GET'])
@login_required
def visualizarFormulario(file):
    path = os.path.join(os.getcwd(), 'SGI_v0/app/static/Word/formularios')
    arquivo = os.path.join(path, file)
    with open(arquivo, "rb") as docx_file:
        result = mammoth.convert_to_html(docx_file)
    html = result.value
    return render_template('formulario.html', html_code=html)



    #####################################################################
    #####################################################################
    ####                    Constatação      ##############################
    #####################################################################

@dimci.route('/registrarConstatacao', methods=['GET','POST'])
@login_required
def registrarConstatacao():

    DOWNLOAD_FOLDER = os.path.join(os.getcwd(), 'SGI_v0/app/static/Excel/Planilhas_Dimci')

    files = os.listdir(DOWNLOAD_FOLDER)
    if request.method == 'POST':
        origem = request.form['origem']
        nconstatacao = request.form['nconstatacao']
        tipo = request.form['tipo']
        inputUO = request.form['inputUO']

        try:
          data_const = datetime.strptime(request.form['data_const'],'%Y-%m-%d')
        except:
          data_const = None

        grandeza = request.form['grandeza']
        laboratorio = request.form['laboratorio']
        responsavel = request.form['responsavel']
        documento_origem = request.form['documento_origem']
        descricao_da_const = request.form['descricao_da_const']
        item_do_enquadramento = request.form['item_do_enquadramento']
        correcao_da_const = request.form['correcao_da_const']
        causa_raiz_da_const = request.form['causa_raiz_da_const']
        acao_corretiva_da_const = request.form['acao_corretiva_da_const']

        try:
           prazo_para_implementacao = datetime.strptime(request.form['prazo_para_implementacao'],'%Y-%m-%d')
        except:
           prazo_para_implementacao = None

        comentario = request.form['comentario']

        try:
           data_da_implementacao = datetime.strptime(request.form['data_da_implementacao'],'%Y-%m-%d')
        except:
           data_da_implementacao = None


        implementado = request.form['implementado']
        repactuacao1 = request.form['repactuacao1']
        repactuacao2 = request.form['repactuacao2']
        repactuacao3 = request.form['repactuacao3']

        if request.form['action'] == 'preencher':

            FOLDER_WORD = os.path.join(os.getcwd(), 'SGI_v0/app/static/Word/templates')
            pathWord = os.path.join(FOLDER_WORD, 'for_constatacao.docx')

            document = MailMerge(pathWord)
            print('PREENCHER FORMULARIO',document.get_merge_fields())
            document.merge(
                tipo=tipo,
                nconstatacao=nconstatacao,
                inputUO=inputUO,
                origem=origem,
                documento_origem=documento_origem,
                descricao_da_const=descricao_da_const,
                item_do_enquadramento=item_do_enquadramento,
                correcao_da_const=correcao_da_const,
                causa_raiz_da_const=causa_raiz_da_const,
                acao_corretiva_da_const=acao_corretiva_da_const,
                prazo_para_implementacao=prazo_para_implementacao.strftime(("%d/%m/%Y")))

            FOLDER_WORD = os.path.join(os.getcwd(), 'SGI_v0/app/static/Word/formularios')
            nconstatacaoString = nconstatacao.replace('/','_')
            nomeArquivo = 'for_constatacao_{}.docx'.format(nconstatacaoString)

            document.write(FOLDER_WORD + '/' + nomeArquivo)

            #arquivo = os.path.join(FOLDER_WORD, nomeArquivo)

            #return send_file(arquivo, mimetype='imagem/png')

        planilha = request.form['planilha']

        pathPlanilha = os.path.join(DOWNLOAD_FOLDER, planilha)
        wb = ''
        try:
           wb = openpyxl.load_workbook(pathPlanilha)
        except:
           flash("Não selecionou a planilha")
           return render_template('registro_constatacao.html', planilhas = files)

        #wb =openpyxl.load_workbook(pathPlanilha)
        ws = wb.get_sheet_by_name(str(wb.worksheets[0].title))
        i=1
        while ws.cell(row=i, column=1).value != None  :
                if ws.cell(row=i, column=1).value == nconstatacao:
                    break
                i +=1
        if request.form['action'] == 'deletar':
            ws.delete_rows(i)
            wb.save(pathPlanilha)
            return redirect(url_for('dimci.consultarConstatacao'))

        ws.cell(row=i, column=1,value=nconstatacao)
        ws.cell(row=i, column=2, value=data_const)
        ws.cell(row=i, column=3, value=inputUO)
        ws.cell(row=i, column=4, value=laboratorio)
        ws.cell(row=i, column=5, value=grandeza)
        ws.cell(row=i, column=6, value=descricao_da_const)
        ws.cell(row=i, column=7, value=tipo)
        ws.cell(row=i, column=8, value='')
        ws.cell(row=i, column=9, value=origem)
        ws.cell(row=i, column=10, value=documento_origem)
        ws.cell(row=i, column=11, value=item_do_enquadramento)
        ws.cell(row=i, column=12, value='')
        ws.cell(row=i, column=13, value=correcao_da_const)
        ws.cell(row=i, column=14, value=causa_raiz_da_const)
        ws.cell(row=i, column=15, value='')
        ws.cell(row=i, column=16, value=acao_corretiva_da_const)
        ws.cell(row=i, column=17, value='')
        ws.cell(row=i, column=18, value='')
        ws.cell(row=i, column=19, value=responsavel)
        ws.cell(row=i, column=20, value=prazo_para_implementacao)
        ws.cell(row=i, column=21, value=comentario)
        ws.cell(row=i, column=22, value=implementado)
        ws.cell(row=i, column=23, value=data_da_implementacao)
        ws.cell(row=i, column=24, value=repactuacao1)
        ws.cell(row=i, column=25, value=repactuacao2)
        ws.cell(row=i, column=26, value=repactuacao3)

        wb.save(pathPlanilha)
        if request.form['action'] == 'editar':
            return redirect(url_for('dimci.consultarConstatacao'))
        else:
            return redirect(url_for('dimci.registrarConstatacao'))

        return redirect(url_for('dimci.registrarConstatacao'))


    return render_template('registro_constatacao.html', planilhas = files)

@dimci.route('/consultarConstatacao', methods=['GET','POST'])
@login_required
def consultarConstatacao():
    DOWNLOAD_FOLDER = os.path.join(os.getcwd(), 'SGI_v0/app/static/Excel/Planilhas_Dimci')
    files = os.listdir(DOWNLOAD_FOLDER)
    if request.method == 'POST':

        planilha = request.form['planilha']
        nconstatacao = request.form['nconstatacao']

        pathPlanilha = os.path.join(DOWNLOAD_FOLDER, planilha)
        wb = ''
        try:
            wb = openpyxl.load_workbook(pathPlanilha)
        except:
           flash("Não selecionou a planilha")
           return render_template('consulta_constatacao.html', planilhas = files)
        ws = wb.get_sheet_by_name(str(wb.worksheets[0].title))

        achou = False
        for i in range(100):
            i+=1
            #print(i,'< --- >' ,nconstatacao ," < - - - >",ws.cell(row=i, column=1).value )
            if ws.cell(row=i, column=1).value == nconstatacao:

                achou = True
                break
        if achou == False:
            flash('N° da constatação não encontrado')
            return render_template('consulta_constatacao.html', planilhas = files)

        nconstatacao = ws.cell(row=i, column=1).value
        data_const = ws.cell(row=i, column=2).value.strftime(("%Y-%m-%d" ))
        inputUO = ws.cell(row=i, column=3).value
        laboratorio = ws.cell(row=i, column=4).value
        grandeza = ws.cell(row=i, column=5).value
        descricao_da_const = ws.cell(row=i, column=6).value
        tipo = ws.cell(row=i, column=7).value
        #ws.cell(row=i, column=8, value='')
        origem = ws.cell(row=i, column=9).value
        documento_origem = ws.cell(row=i, column=10).value
        item_do_enquadramento = ws.cell(row=i, column=11).value
        #ws.cell(row=i, column=12, value='')
        correcao_da_const = ws.cell(row=i, column=13).value
        causa_raiz_da_const = ws.cell(row=i, column=14).value
        #ws.cell(row=i, column=15, value='')
        acao_corretiva_da_const = ws.cell(row=i, column=16).value
        #ws.cell(row=i, column=17, value='')
        #ws.cell(row=i, column=18, value='')
        responsavel =  ws.cell(row=i, column=19).value
        prazo_para_implementacao = ws.cell(row=i, column=20).value.strftime(("%Y-%m-%d" ))
        comentario = ws.cell(row=i, column=21).value
        implementado =ws.cell(row=i, column=22).value
        data_da_implementacao = ws.cell(row=i, column=23).value.strftime(("%Y-%m-%d" ))
        repactuacao1 = ws.cell(row=i, column=24).value
        repactuacao2 = ws.cell(row=i, column=25).value
        repactuacao3 = ws.cell(row=i, column=26).value

        return render_template('mostrar_constatacao.html',  planilha = planilha, planilhas = files, nconstatacao = nconstatacao,data_const=data_const,inputUO=inputUO,
                               laboratorio=laboratorio,grandeza=grandeza, descricao_da_const=descricao_da_const,tipo=tipo,
                               origem=origem,documento_origem=documento_origem,item_do_enquadramento=item_do_enquadramento,correcao_da_const=correcao_da_const,
                               causa_raiz_da_const=causa_raiz_da_const,acao_corretiva_da_const=acao_corretiva_da_const,responsavel=responsavel,
                               prazo_para_implementacao=prazo_para_implementacao,comentario=comentario,implementado=implementado,data_da_implementacao=data_da_implementacao,
                               repactuacao1=repactuacao1,repactuacao2=repactuacao2,repactuacao3=repactuacao3)


    files = os.listdir(DOWNLOAD_FOLDER)
    return render_template('consulta_constatacao.html', planilhas = files)